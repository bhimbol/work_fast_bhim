from django.shortcuts import redirect, render, HttpResponseRedirect
from django.http import HttpResponse
from django.conf import settings
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from datetime import datetime
import requests
import os
import json
import base64
import shutil
import ast
import pandas as pd
import os.path
import os
import xlrd
from openpyxl import load_workbook

def asn_manager_dashboard(request):
    return render(request, 'asn_manager_dashboard.html')

def login(request):
    return 0

def logout(request):
    try:
        authenticated = request.session.get('google_authenticated', False)
        if authenticated:
            google_credentials_json = request.session.get('google_credentials')
            credentials = Credentials.from_authorized_user_info(json.loads(google_credentials_json))
            revoke_url = "https://accounts.google.com/o/oauth2/revoke"
            revoke_params = {"token": credentials.token}
            requests.post(revoke_url, params=revoke_params)
            del request.session['user_info']
            del request.session['google_credentials']
            request.session['google_authenticated'] = False
    except KeyError as e:
        return HttpResponse(f"KeyError: {e}")
    except Exception as e:
        return HttpResponse(f"An Exception occurred: {e}")
    return redirect('asn_manager:asn_manager_dashboard')

def initiate_google_auth(request):
    credentials_path = os.path.join(settings.BASE_DIR, 'common_files', 'keyMe.json')
    with open(credentials_path, 'r') as file:
        credentials_data = json.load(file)
    flow = InstalledAppFlow.from_client_config(
        credentials_data,
        scopes=['https://www.googleapis.com/auth/gmail.readonly'],
        redirect_uri='https://airbytes.pythonanywhere.com/asn_manager/auth_callback'
    )
    authorization_url, state = flow.authorization_url(prompt="consent")
    request.session['google_auth_state'] = state
    return HttpResponseRedirect(authorization_url)

def get_google_credentials(authorization_code):
    credentials_path = os.path.join(settings.BASE_DIR, 'common_files', 'keyMe.json')
    with open(credentials_path, 'r') as file:
        credentials_data = json.load(file)
    flow = InstalledAppFlow.from_client_config(
        credentials_data,
        scopes=['https://www.googleapis.com/auth/gmail.readonly'],
        redirect_uri='https://airbytes.pythonanywhere.com/asn_manager/auth_callback'
    )
    flow.fetch_token(code=authorization_code)
    return flow.credentials

def get_user_info(credentials):
    gmail_service = build('gmail', 'v1', credentials=credentials)
    user_info = gmail_service.users().getProfile(userId='me').execute()
    return user_info

def auth_callback(request):
    authorization_code = request.GET.get('code')
    received_state = request.GET.get('state')
    expected_state = request.session.pop('google_auth_state', None)

    if received_state != expected_state:
        return render(request, 'csrf_error.html', {'received_state': received_state, 'expected_state': expected_state})

    try:
        credentials = get_google_credentials(authorization_code)
        user_info = get_user_info(credentials)
        request.session['user_info'] = user_info
        request.session['google_credentials'] = credentials.to_json()
        request.session['google_authenticated'] = True
    except KeyError as e:
        return HttpResponse(f"KeyError: {e}")
    except Exception as e:
        return HttpResponse(f"An Exception occurred: {e}")
    return redirect('asn_manager:asn_manager_dashboard')

def get_gmail_messages(request, keywords=None, rdd=None, sender=None, has_attachment=True):
    try:
        if 'google_credentials' in request.session:
            google_credentials_json = request.session.get('google_credentials')
            credentials = Credentials.from_authorized_user_info(json.loads(google_credentials_json))
            gmail_service = build('gmail', 'v1', credentials=credentials)

            query_parts = []
            if has_attachment:
                query_parts.append('has:attachment')
            if sender:
                query_parts.append(f'from:{sender}')
            if keywords:
                for keyword in keywords:
                    query_parts.append(f'subject:{keyword}')
            query = ' '.join(query_parts)

            messages = gmail_service.users().messages().list(userId='me', q=query).execute()
            message_list = []
            for message in messages.get('messages', []):
                message_details = gmail_service.users().messages().get(userId='me', id=message['id']).execute()
                subject = next((header['value'] for header in message_details['payload']['headers'] if header['name'] == 'Subject'), 'N/A')
                attachments = []
                if 'parts' in message_details['payload']:
                    for part in message_details['payload']['parts']:
                        if 'data' in part['body']:
                            data = part['body']['data']
                        elif 'attachmentId' in part['body']:
                            att_id = part['body']['attachmentId']
                            att = gmail_service.users().messages().attachments().get(userId='me', messageId=message['id'], id=att_id).execute()
                            data = att['data']
                        file_data = base64.urlsafe_b64decode(data.encode('UTF-8'))
                        attachments.append({'filename': part['filename'], 'data': file_data})
                message_list.append({'id': message['id'], 'subject': subject, 'attachments': attachments})
            return message_list
    except KeyError as e:
        return HttpResponse(f"KeyError: {e}")
    except Exception as e:
        return HttpResponse(f"An Exception occurred: {e}")
    return []

def view_asn_list(request):
    if 'google_credentials' in request.session:
        branch = request.POST.get('branch', '')
        rdd = request.POST.get('rdd', '')
        rdd_datetime = datetime.strptime(rdd, '%Y-%m-%d')
        formatted_rdd_datetime = rdd_datetime.strftime('%d %b %Y')

        sender = request.POST.get('sender', '')
        messages = []
        messages_count = 0
        if request.method == 'POST':
            try:
                messages = get_gmail_messages(request, keywords=[branch, formatted_rdd_datetime], sender=sender, has_attachment=True)
                if messages:
                    messages_count = len(messages)
                    request.session['message_count'] = messages_count
            except KeyError as e:
                return HttpResponse(f"KeyError: {e}")
            except Exception as e:
                return HttpResponse(f"An Exception occurred: {e}")
    return render(request, 'asn_manager_dashboard.html', {'branch': branch, 'rdd': rdd, 'sender': sender, 'messages': messages})


def download_selected_asn_list(request):
    if request.method == 'POST':
        selected_messages_list = request.POST.getlist('selected_messages')
        include_dat = request.POST.get('include_dat')
        save_dir = os.path.join(settings.BASE_DIR, 'downloaded_asn_files', '')
        list = []
        if os.path.exists(save_dir):
            shutil.rmtree(save_dir)
        os.makedirs(save_dir)

        try:
            for message_str in selected_messages_list:
                message = ast.literal_eval(message_str)
                attachments = message.get('attachments', [])
                for attachment in attachments:
                    filename = attachment.get('filename', '')
                    file_path = os.path.join(save_dir, filename)
                    data = attachment.get('data', '')
                    if '.xls' in filename or (include_dat and '.DAT' in filename):
                        with open(file_path, 'wb') as file:
                            file.write(data)
                list.append(message)
            #return render(request, 'download_selected_asn_list.html', {'branch': branch, 'rdd': rdd, 'sender': sender, 'messages': messages})
            return render(request, 'download_selected_asn_list.html',{'messages':list})
        except KeyError as e:
            return HttpResponse(f"KeyError: {e}")
        except Exception as e:
            return HttpResponse(f"An Exception occurred: {e}")
    else:
        return HttpResponse("Method not allowed. Please use a POST request.")

def append_contents_to_existing_final(existing_final_file_path, list_of_files):
    final_workbook = load_workbook(existing_final_file_path)
    final_sheet = final_workbook['DATA']
    final_sheet.delete_rows(1, final_sheet.max_row)
    for file_path in list_of_files:
        source_workbook = xlrd.open_workbook(file_path)
        source_sheet = source_workbook.sheet_by_index(0)
        for row_index in range(source_sheet.nrows):
            new_row = source_sheet.row_values(row_index)
            final_sheet.append(new_row)
    data_column_i_values = [final_sheet.cell(row=row_index, column=9).value for row_index in range(1, final_sheet.max_row + 1)]
    comparison_sheet = final_workbook['COMPARISON']
    for row in comparison_sheet.iter_rows(min_row=2, max_row=comparison_sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = None
    unique_values = list(set(data_column_i_values))
    unique_values = [value for value in unique_values if value is not None and pd.notna(value)]
    start_row = 2
    for value in unique_values:
        comparison_sheet.cell(row=start_row, column=1, value=value)
        start_row += 1
    final_workbook.save(existing_final_file_path)
    with open(existing_final_file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={existing_final_file_path}'
    return response

def consolidate_asn(request):
    existing_file_name = 'summary.xlsx'
    existing_file_path = os.path.join(settings.BASE_DIR, 'common_files', existing_file_name)
    asn_list = os.path.join(settings.BASE_DIR, 'downloaded_asn_files')
    asn_list_file_paths = [os.path.join(asn_list, filename) for filename in os.listdir(asn_list) if filename.endswith('.xls')]
    response = append_contents_to_existing_final(existing_file_path, asn_list_file_paths)
    #if response.status_code == 200:
    #    return redirect('asn_manager:asn_manager_dashboard')
    return response